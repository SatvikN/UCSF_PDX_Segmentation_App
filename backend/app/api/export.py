import os
import io
import zipfile
from typing import Literal

import numpy as np
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from PIL import Image
from scipy.io import savemat

from app.services.dicom import list_dicom_files
from app.services.metadata import read_study_info, read_spacing_and_thickness_mm
from app.services.storage import get_study_subdir, get_study_dicom_source_dir
from app.services.images import ensure_png_slices
from app.services.volume import compute_raw_areas


router = APIRouter(prefix="/export", tags=["export"])


@router.get("/{study_id}/images.zip")
async def export_images(
    study_id: str,
    kind: Literal["overlays", "masks", "pngs"] = Query("overlays", description="Export overlays, masks, or PNG images"),
    prefix: str | None = Query(None, description="Optional filename prefix inside the ZIP"),
):
    if kind == "pngs":
        ensure_png_slices(study_id)
        target_dir = get_study_subdir(study_id, "png")
    else:
        target_dir = get_study_subdir(study_id, "overlays" if kind == "overlays" else "masks")
    if not os.path.isdir(target_dir):
        raise HTTPException(status_code=404, detail=f"{kind} not found")
    files = [f for f in sorted(os.listdir(target_dir)) if f.lower().endswith('.png')]
    if not files:
        raise HTTPException(status_code=404, detail=f"no {kind} to export")

    # Map indices to original DICOM-derived base names when available
    dicom_dir = get_study_dicom_source_dir(study_id)
    dcm_files = list_dicom_files(dicom_dir)
    dcm_bases = [os.path.splitext(n)[0] for n in dcm_files]

    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for i, name in enumerate(files):
            # Prefer original DICOM base name if lengths align; else fall back to existing name
            base_without_ext = dcm_bases[i] if i < len(dcm_bases) else os.path.splitext(name)[0]
            arc_name = f"{base_without_ext}.png"
            if prefix:
                arc_name = f"{prefix}_{arc_name}"
            zf.write(os.path.join(target_dir, name), arcname=arc_name)
    mem.seek(0)
    filename = "images.zip" if kind == "pngs" else f"{kind}.zip"
    if prefix:
        filename = f"{prefix}_" + filename
    return StreamingResponse(mem, media_type="application/zip", headers={
        "Content-Disposition": f"attachment; filename={filename}"
    })


@router.get("/{study_id}/volumes.xlsx")
async def export_volumes_excel(study_id: str, prefix: str | None = Query(None)):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Prepare data
    dicom_dir = get_study_dicom_source_dir(study_id)
    dcm_files = list_dicom_files(dicom_dir)
    if not dcm_files:
        raise HTTPException(status_code=404, detail="no slices found")
    masks_dir = get_study_subdir(study_id, "masks")
    mask_pngs = [os.path.join(masks_dir, f"{i+1}.png") for i in range(len(dcm_files))]
    # Compute raw pixel counts per slice from masks
    raw_counts = []
    for p in mask_pngs:
        if not os.path.exists(p):
            raw_counts.append(0)
            continue
        arr = np.array(Image.open(p).convert('L'))
        raw_counts.append(int((arr > 127).sum()))
    pixel_spacing_mm, thickness_mm = read_spacing_and_thickness_mm(study_id)
    meta = read_study_info(study_id)

    # Build workbook with formulas
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volumes"

    # Metadata block
    ws.append(["Study ID", study_id])
    ws.append(["Patient ID", meta.get("patient_id", "")])
    ws.append(["Study Date", meta.get("study_date", "")])
    ws.append(["Pixel Spacing mm", f"{pixel_spacing_mm[0]}", f"{pixel_spacing_mm[1]}"])
    px_row = ws.max_row
    ws.append(["Slice Thickness mm", thickness_mm])
    th_row = ws.max_row
    ws.append(["Spacing Between Slices (mm)", meta.get("spacing_between_slices_mm", "")])
    ws.append(["Image Height (px)", meta.get("height", "")])
    ws.append(["Image Width (px)", meta.get("width", "")])
    # Two blank rows before table
    ws.append(["", "", ""])  # Empty row with 3 columns
    ws.append(["", "", ""])  # Empty row with 3 columns

    # Header for per-slice table
    start_row = ws.max_row + 1
    ws.cell(row=start_row, column=1, value="Slice (DICOM)")
    ws.cell(row=start_row, column=2, value="Raw area (pixels)")
    ws.cell(row=start_row, column=3, value="Scaled area (cc)")

    # Constants cells for formulas (pixel spacing and thickness)
    pxr, pxc = px_row, 2  # pixel spacing x at column B
    pyr, pyc = px_row, 3  # pixel spacing y at column C
    thr, thc = th_row, 2  # thickness at column B

    for i, (name, raw) in enumerate(zip(dcm_files, raw_counts), start=1):
        r = start_row + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=raw)
        # Formula: raw * thickness_mm * px * py / 1000
        ws.cell(row=r, column=3, value=f"=B{r}*$B{thr}*$B{pxr}*$C{pyr}/1000")

    # Total row
    end_row = start_row + len(dcm_files)
    ws.cell(row=end_row + 1, column=1, value="Total volume (cc)")
    ws.cell(row=end_row + 1, column=3, value=f"=SUM(C{start_row+1}:C{end_row})")

    # Autosize some columns
    for col in range(1, 4):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 22

    # Stream workbook
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    out_name = "volumes.xlsx"
    if prefix:
        out_name = f"{prefix}_" + out_name
    return StreamingResponse(mem, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f"attachment; filename={out_name}"
    })


@router.get("/{study_id}/images.npz")
async def export_images_npz(study_id: str, prefix: str | None = Query(None)):
    ensure_png_slices(study_id)
    png_dir = get_study_subdir(study_id, "png")
    files = [f for f in sorted(os.listdir(png_dir)) if f.lower().endswith('.png')]
    if not files:
        raise HTTPException(status_code=404, detail="no images available")
    arrs = []
    for name in files:
        p = os.path.join(png_dir, name)
        a = np.array(Image.open(p).convert('L'))
        arrs.append(a.astype(np.uint8))
    vol = np.stack(arrs, axis=0)  # N,H,W
    mem = io.BytesIO()
    np.savez_compressed(mem, images=vol)
    mem.seek(0)
    out_name = "images.npz"
    if prefix:
        out_name = f"{prefix}_" + out_name
    return StreamingResponse(mem, media_type="application/octet-stream", headers={
        "Content-Disposition": f"attachment; filename={out_name}"
    })


@router.get("/{study_id}/masks")
async def export_masks(study_id: str, format: str = Query("npz", description="Export format: npz, mat"), prefix: str | None = Query(None)):
    """
    Export masks in various formats.
    Supports NPZ (NumPy) and MAT (MATLAB) formats.
    """
    masks_dir = get_study_subdir(study_id, "masks")
    files = [f for f in sorted(os.listdir(masks_dir)) if f.lower().endswith('.png')]
    if not files:
        raise HTTPException(status_code=404, detail="no masks available")
    
    # Load all mask images and convert to binary arrays
    mask_arrays = []
    for name in files:
        p = os.path.join(masks_dir, name)
        img = Image.open(p).convert('L')
        # Convert to binary mask (0 or 1)
        mask = (np.array(img) > 127).astype(np.uint8)
        mask_arrays.append(mask)
    
    # Stack all masks into a 3D array (num_slices, height, width)
    masks_volume = np.stack(mask_arrays, axis=0)
    
    # Create file in memory based on format
    mem = io.BytesIO()
    
    if format.lower() == "mat":
        # MATLAB .mat format
        savemat(mem, {'masks': masks_volume}, do_compression=True)
        file_extension = "mat"
        media_type = "application/octet-stream"
    else:
        # NumPy .npz format (default)
        np.savez_compressed(mem, masks=masks_volume)
        file_extension = "npz"
        media_type = "application/octet-stream"
    
    mem.seek(0)
    
    # Generate filename
    out_name = f"masks.{file_extension}"
    if prefix:
        out_name = f"{prefix}_{out_name}"
    
    return StreamingResponse(
        mem, 
        media_type=media_type, 
        headers={
            "Content-Disposition": f"attachment; filename={out_name}"
        }
    )


